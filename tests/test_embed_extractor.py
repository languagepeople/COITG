"""
tests/test_embed_extractor.py
==============================
Unit tests for embed_extractor.py.

These tests do NOT require a browser or network access; they cover the
pure-Python helpers that can be exercised offline.
"""

import csv
import os
import sys
import tempfile
import uuid as _uuid_module
from unittest.mock import patch

import pytest

# Ensure the project root is on the path so we can import the script directly
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import embed_extractor as ee


# ---------------------------------------------------------------------------
# extract_youtube_id
# ---------------------------------------------------------------------------


class TestExtractYoutubeId:
    def test_standard_watch_url(self):
        assert ee.extract_youtube_id("https://www.youtube.com/watch?v=dQw4w9WgXcQ") == "dQw4w9WgXcQ"

    def test_short_url(self):
        assert ee.extract_youtube_id("https://youtu.be/dQw4w9WgXcQ") == "dQw4w9WgXcQ"

    def test_shorts_url(self):
        assert ee.extract_youtube_id("https://www.youtube.com/shorts/abc12345678") == "abc12345678"

    def test_embed_url(self):
        assert ee.extract_youtube_id("https://www.youtube.com/embed/dQw4w9WgXcQ") == "dQw4w9WgXcQ"

    def test_url_with_extra_params(self):
        assert (
            ee.extract_youtube_id(
                "https://www.youtube.com/watch?v=dQw4w9WgXcQ&t=42s&list=PLxxx"
            )
            == "dQw4w9WgXcQ"
        )

    def test_non_youtube_url_returns_none(self):
        assert ee.extract_youtube_id("https://www.example.com/video/12345") is None

    def test_empty_string_returns_none(self):
        assert ee.extract_youtube_id("") is None


# ---------------------------------------------------------------------------
# build_youtube_embed
# ---------------------------------------------------------------------------


class TestBuildYoutubeEmbed:
    def test_contains_iframe(self):
        html = ee.build_youtube_embed("dQw4w9WgXcQ")
        assert html.startswith("<iframe")
        assert html.endswith("></iframe>")

    def test_contains_video_id(self):
        html = ee.build_youtube_embed("dQw4w9WgXcQ")
        assert "dQw4w9WgXcQ" in html

    def test_contains_embed_src(self):
        html = ee.build_youtube_embed("dQw4w9WgXcQ")
        assert 'src="https://www.youtube.com/embed/dQw4w9WgXcQ"' in html

    def test_allowfullscreen(self):
        html = ee.build_youtube_embed("dQw4w9WgXcQ")
        assert "allowfullscreen" in html


# ---------------------------------------------------------------------------
# _parse_iso8601_duration
# ---------------------------------------------------------------------------


class TestParseIso8601Duration:
    def test_minutes_seconds(self):
        assert ee._parse_iso8601_duration("PT4M33S") == "4:33"

    def test_hours_minutes_seconds(self):
        assert ee._parse_iso8601_duration("PT1H2M3S") == "1:02:03"

    def test_seconds_only(self):
        assert ee._parse_iso8601_duration("PT45S") == "0:45"

    def test_minutes_only(self):
        assert ee._parse_iso8601_duration("PT10M") == "10:00"

    def test_hours_only(self):
        assert ee._parse_iso8601_duration("PT2H") == "2:00:00"

    def test_days_and_time(self):
        assert ee._parse_iso8601_duration("P1DT1H") == "25:00:00"

    def test_unrecognised_returns_original(self):
        assert ee._parse_iso8601_duration("NOT_VALID") == "NOT_VALID"


# ---------------------------------------------------------------------------
# get_video_duration_from_html
# ---------------------------------------------------------------------------


class TestGetVideoDurationFromHtml:
    def test_standard_meta_tag(self):
        html = '<meta itemprop="duration" content="PT4M33S">'
        assert ee.get_video_duration_from_html(html) == "4:33"

    def test_single_quote_meta_tag(self):
        html = "<meta itemprop='duration' content='PT1H2M3S'>"
        assert ee.get_video_duration_from_html(html) == "1:02:03"

    def test_content_before_itemprop(self):
        html = '<meta content="PT10M" itemprop="duration">'
        assert ee.get_video_duration_from_html(html) == "10:00"

    def test_missing_duration_returns_none(self):
        html = "<html><body>No duration here</body></html>"
        assert ee.get_video_duration_from_html(html) is None


# ---------------------------------------------------------------------------
# _get_embed  (no driver → programmatic fallback, backward compat)
# ---------------------------------------------------------------------------


class TestGetEmbedNoDriver:
    def test_youtube_url_returns_iframe(self):
        code = ee._get_embed(None, "https://www.youtube.com/watch?v=dQw4w9WgXcQ")
        assert code is not None
        assert "<iframe" in code
        assert "dQw4w9WgXcQ" in code

    def test_non_video_url_returns_none(self):
        code = ee._get_embed(None, "https://www.example.com/page")
        assert code is None


# ---------------------------------------------------------------------------
# _get_embed_and_duration  (no driver → fallback, duration via mocked network)
# ---------------------------------------------------------------------------


class TestGetEmbedAndDurationNoDriver:
    def test_returns_embed_and_duration(self, monkeypatch):
        monkeypatch.setattr(ee, "fetch_youtube_duration", lambda url: "4:33")
        embed, duration = ee._get_embed_and_duration(
            None, "https://www.youtube.com/watch?v=dQw4w9WgXcQ"
        )
        assert embed is not None
        assert "<iframe" in embed
        assert duration == "4:33"

    def test_non_youtube_no_embed_no_duration(self):
        embed, duration = ee._get_embed_and_duration(
            None, "https://www.example.com/page"
        )
        assert embed is None
        assert duration is None


# ---------------------------------------------------------------------------
# Fixtures: suppress browser and network calls in integration tests
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def no_browser_no_network(monkeypatch):
    """Prevent any test from spawning a real browser or making network calls."""
    monkeypatch.setattr(ee, "_create_driver", lambda headless=True: None)
    monkeypatch.setattr(ee, "fetch_youtube_duration", lambda url: None)


# ---------------------------------------------------------------------------
# process_csv  (end-to-end with a temp file, no browser)
# ---------------------------------------------------------------------------


class TestProcessCsv:
    def _make_csv(self, rows, headers=None):
        """Write rows to a temp CSV and return its path."""
        fd, path = tempfile.mkstemp(suffix=".csv")
        os.close(fd)
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            if headers:
                writer.writerow(headers)
            writer.writerows(rows)
        return path

    def test_embed_code_written_to_new_column(self):
        path = self._make_csv(
            [["https://www.youtube.com/watch?v=dQw4w9WgXcQ"]],
            headers=["Video URL"],
        )
        try:
            ee.process_csv(path, "1", "2", headless=True)

            with open(path, newline="", encoding="utf-8-sig") as f:
                rows = list(csv.reader(f))

            assert len(rows) == 2  # header + 1 data row
            assert "<iframe" in rows[1][1]
            assert "dQw4w9WgXcQ" in rows[1][1]
        finally:
            os.unlink(path)

    def test_embed_code_by_header_name(self):
        path = self._make_csv(
            [["https://youtu.be/dQw4w9WgXcQ", ""]],
            headers=["URL", "Embed"],
        )
        try:
            ee.process_csv(path, "URL", "Embed", headless=True)

            with open(path, newline="", encoding="utf-8-sig") as f:
                rows = list(csv.reader(f))

            assert "<iframe" in rows[1][1]
        finally:
            os.unlink(path)

    def test_empty_url_rows_skipped(self):
        path = self._make_csv(
            [[""], ["https://www.youtube.com/watch?v=dQw4w9WgXcQ"]],
            headers=["URL"],
        )
        try:
            ee.process_csv(path, "1", "2", headless=True)

            with open(path, newline="", encoding="utf-8-sig") as f:
                rows = list(csv.reader(f))

            # Row with empty URL should have no embed code
            assert rows[1][1] == ""
            # Row with YouTube URL should have embed code
            assert "<iframe" in rows[2][1]
        finally:
            os.unlink(path)

    def test_non_youtube_url_no_embed(self):
        path = self._make_csv(
            [["https://www.example.com/not-a-video"]],
            headers=["URL"],
        )
        try:
            ee.process_csv(path, "1", "2", headless=True)

            with open(path, newline="", encoding="utf-8-sig") as f:
                rows = list(csv.reader(f))

            assert rows[1][1] == ""
        finally:
            os.unlink(path)

    def test_duration_written_to_third_column(self, monkeypatch):
        monkeypatch.setattr(ee, "fetch_youtube_duration", lambda url: "3:45")
        path = self._make_csv(
            [["https://www.youtube.com/watch?v=dQw4w9WgXcQ"]],
            headers=["URL"],
        )
        try:
            ee.process_csv(path, "1", "2", headless=True, duration_col="3")

            with open(path, newline="", encoding="utf-8-sig") as f:
                rows = list(csv.reader(f))

            assert "<iframe" in rows[1][1]
            assert rows[1][2] == "3:45"
        finally:
            os.unlink(path)

    def test_item_id_written_to_column(self):
        path = self._make_csv(
            [["https://www.youtube.com/watch?v=dQw4w9WgXcQ"]],
            headers=["URL"],
        )
        try:
            ee.process_csv(path, "1", "2", headless=True, id_col="3")

            with open(path, newline="", encoding="utf-8-sig") as f:
                rows = list(csv.reader(f))

            item_id = rows[1][2]
            assert item_id  # non-empty
            # Should be a valid UUID
            _uuid_module.UUID(item_id)  # raises ValueError if not a valid UUID
        finally:
            os.unlink(path)

    def test_item_id_not_written_for_empty_url(self):
        path = self._make_csv(
            [[""], ["https://www.youtube.com/watch?v=dQw4w9WgXcQ"]],
            headers=["URL"],
        )
        try:
            ee.process_csv(path, "1", "2", headless=True, id_col="3")

            with open(path, newline="", encoding="utf-8-sig") as f:
                rows = list(csv.reader(f))

            # Empty-URL row should have no ID
            assert rows[1][2] == ""
            # YouTube row should have an ID
            assert rows[2][2]
        finally:
            os.unlink(path)


# ---------------------------------------------------------------------------
# process_excel  (end-to-end with a temp .xlsx, no browser)
# ---------------------------------------------------------------------------


class TestProcessExcel:
    def _make_xlsx(self, rows, headers=None):
        import openpyxl

        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb = openpyxl.Workbook()
        ws = wb.active
        if headers:
            ws.append(headers)
        for row in rows:
            ws.append(row)
        wb.save(path)
        return path

    def test_embed_code_written_to_column(self):
        import openpyxl

        path = self._make_xlsx(
            [["https://www.youtube.com/watch?v=dQw4w9WgXcQ"]],
            headers=["Video URL"],
        )
        try:
            ee.process_excel(path, "1", "2", headless=True)

            wb = openpyxl.load_workbook(path)
            ws = wb.active
            embed = ws.cell(row=2, column=2).value
            assert embed is not None
            assert "<iframe" in embed
            assert "dQw4w9WgXcQ" in embed
        finally:
            os.unlink(path)

    def test_embed_code_by_header_name(self):
        import openpyxl

        path = self._make_xlsx(
            [["https://youtu.be/dQw4w9WgXcQ", ""]],
            headers=["URL", "Embed Code"],
        )
        try:
            ee.process_excel(path, "URL", "Embed Code", headless=True)

            wb = openpyxl.load_workbook(path)
            ws = wb.active
            embed = ws.cell(row=2, column=2).value
            assert embed is not None
            assert "<iframe" in embed
        finally:
            os.unlink(path)

    def test_duration_written_to_third_column(self, monkeypatch):
        import openpyxl

        monkeypatch.setattr(ee, "fetch_youtube_duration", lambda url: "4:33")
        path = self._make_xlsx(
            [["https://www.youtube.com/watch?v=dQw4w9WgXcQ"]],
            headers=["URL"],
        )
        try:
            ee.process_excel(path, "1", "2", headless=True, duration_col="3")

            wb = openpyxl.load_workbook(path)
            ws = wb.active
            embed = ws.cell(row=2, column=2).value
            duration = ws.cell(row=2, column=3).value
            assert embed is not None
            assert "<iframe" in embed
            assert duration == "4:33"
        finally:
            os.unlink(path)

    def test_item_id_written_to_column(self):
        import openpyxl

        path = self._make_xlsx(
            [["https://www.youtube.com/watch?v=dQw4w9WgXcQ"]],
            headers=["URL"],
        )
        try:
            ee.process_excel(path, "1", "2", headless=True, id_col="3")

            wb = openpyxl.load_workbook(path)
            ws = wb.active
            item_id = ws.cell(row=2, column=3).value
            assert item_id  # non-empty
            # Should be a valid UUID
            _uuid_module.UUID(item_id)  # raises ValueError if not a valid UUID
        finally:
            os.unlink(path)
