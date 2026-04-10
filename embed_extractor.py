#!/usr/bin/env python3
"""
embed_extractor.py
==================
Reads a spreadsheet (Excel .xlsx/.xls or CSV), follows URLs found in a
specified column, navigates each URL in a headless browser, clicks the
Share → Embed buttons, and writes the resulting <iframe> HTML into a
configurable output column on the same row.  Optionally also fetches
the video duration and writes it into a second output column.

YouTube is the primary target.  For YouTube URLs a fast programmatic
fallback is also included so the script still works even when browser
automation is blocked or the page layout has changed.

Default columns match the COITG course-content spreadsheet:
    --url-col  F   (column containing the YouTube / video links)
    --embed-col O  (column to receive the <iframe> embed HTML)
    --duration-col N  (column to receive the video duration, e.g. "4:33")
    --id-col P  (column to receive a generated UUID for each row)

Usage
-----
    python embed_extractor.py <spreadsheet> [options]

Examples
--------
    # Use default columns F (URL), O (embed), N (duration), P (UUID)
    python embed_extractor.py "Course Content.xlsx"

    # Override any column by number, letter, or header name
    python embed_extractor.py videos.xlsx --url-col 2 --embed-col 3 \\
        --duration-col 4 --id-col 5

    # URL column by header name, output to named columns
    python embed_extractor.py videos.csv --url-col "Video URL" \\
        --embed-col "Embed Code" --duration-col "Duration" --id-col "Item ID"

    # Keep the browser window visible (non-headless) for debugging
    python embed_extractor.py videos.xlsx --no-headless
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import time
import uuid
from urllib.error import URLError
from urllib.parse import urlparse
from urllib.request import Request, urlopen

# ---------------------------------------------------------------------------
# Optional dependency: selenium  (installed via requirements.txt)
# ---------------------------------------------------------------------------
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options as ChromeOptions
    from selenium.webdriver.chrome.service import Service as ChromeService
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait

    try:
        from webdriver_manager.chrome import ChromeDriverManager

        _WDM_AVAILABLE = True
    except ImportError:
        _WDM_AVAILABLE = False

    _SELENIUM_AVAILABLE = True
except ImportError:
    _SELENIUM_AVAILABLE = False

# ---------------------------------------------------------------------------
# Optional dependency: openpyxl  (installed via requirements.txt)
# ---------------------------------------------------------------------------
try:
    import openpyxl

    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


# ---------------------------------------------------------------------------
# YouTube helpers
# ---------------------------------------------------------------------------

_YT_PATTERNS = [
    r"(?:youtube\.com/watch\?(?:.*&)?v=|youtu\.be/)([A-Za-z0-9_-]{11})",
    r"youtube\.com/embed/([A-Za-z0-9_-]{11})",
    r"youtube\.com/shorts/([A-Za-z0-9_-]{11})",
]

_YT_HOSTNAMES = frozenset(
    {"www.youtube.com", "youtube.com", "m.youtube.com", "youtu.be"}
)


def _is_youtube_url(url: str) -> bool:
    """Return True only when *url*'s hostname is a known YouTube domain."""
    try:
        hostname = urlparse(url).hostname or ""
        return hostname in _YT_HOSTNAMES
    except Exception:
        return False


def extract_youtube_id(url: str) -> str | None:
    """Return the 11-character YouTube video ID from *url*, or None."""
    for pattern in _YT_PATTERNS:
        m = re.search(pattern, url)
        if m:
            return m.group(1)
    return None


def build_youtube_embed(video_id: str) -> str:
    """Return the standard YouTube <iframe> embed HTML for *video_id*."""
    return (
        f'<iframe width="560" height="315" '
        f'src="https://www.youtube.com/embed/{video_id}" '
        f'title="YouTube video player" '
        f'frameborder="0" '
        f'allow="accelerometer; autoplay; clipboard-write; encrypted-media; '
        f'gyroscope; picture-in-picture; web-share" '
        f'referrerpolicy="strict-origin-when-cross-origin" '
        f'allowfullscreen></iframe>'
    )


def _parse_iso8601_duration(iso: str) -> str:
    """
    Convert an ISO 8601 duration string to a human-readable time string.

    Examples:
        PT4M33S  → "4:33"
        PT1H2M3S → "1:02:03"
        PT45S    → "0:45"
    """
    m = re.fullmatch(
        r"P(?:(\d+)D)?T(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?",
        iso.strip(),
    )
    if not m:
        return iso  # return as-is if unrecognised

    days = int(m.group(1) or 0)
    hours = int(m.group(2) or 0) + days * 24
    minutes = int(m.group(3) or 0)
    seconds = int(m.group(4) or 0)

    if hours:
        return f"{hours}:{minutes:02d}:{seconds:02d}"
    return f"{minutes}:{seconds:02d}"


def get_video_duration_from_html(html: str) -> str | None:
    """
    Parse an ISO 8601 duration from YouTube page HTML and return it as a
    human-readable string (e.g. "4:33"), or None if not found.
    """
    # Primary pattern: itemprop= before content=
    m = re.search(r'itemprop=["\']duration["\'][^>]*content=["\']([^"\']+)["\']', html)
    if not m:
        # Alternate pattern: content= before itemprop=
        m = re.search(r'content=["\']([^"\']+)["\'][^>]*itemprop=["\']duration["\']', html)
    if m:
        candidate = m.group(1)
        if candidate.startswith("P"):
            return _parse_iso8601_duration(candidate)
    return None


def fetch_youtube_duration(url: str) -> str | None:
    """
    Fetch the YouTube page for *url* and return the video duration as a
    human-readable string (e.g. "4:33"), or None on failure.

    This is the network-based fallback used when a browser is unavailable.
    """
    try:
        req = Request(
            url,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                )
            },
        )
        with urlopen(req, timeout=15) as resp:
            html = resp.read().decode("utf-8", errors="replace")
        return get_video_duration_from_html(html)
    except (URLError, OSError):
        return None


def get_duration_via_browser(driver: "webdriver.Chrome", url: str) -> str | None:
    """
    Extract video duration from a page already loaded in *driver*.

    Tries to read the <meta itemprop="duration"> tag from the page source,
    then falls back to network fetch via urllib.
    """
    try:
        page_source = driver.page_source
        duration = get_video_duration_from_html(page_source)
        if duration:
            return duration
    except Exception:
        pass

    # Network fallback (page may still be accessible)
    return fetch_youtube_duration(url)


# ---------------------------------------------------------------------------
# Browser automation
# ---------------------------------------------------------------------------


def _create_driver(headless: bool = True) -> "webdriver.Chrome | None":
    """
    Create and return a Chrome WebDriver instance, or **None** if Chrome /
    ChromeDriver is not available in this environment.

    A *None* return value lets the callers fall back to programmatic embed
    code generation (e.g. for YouTube URLs) without crashing.
    """
    if not _SELENIUM_AVAILABLE:
        return None

    options = ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )

    try:
        if _WDM_AVAILABLE:
            service = ChromeService(ChromeDriverManager().install())
            return webdriver.Chrome(service=service, options=options)
        return webdriver.Chrome(options=options)
    except Exception as exc:
        print(
            f"Warning: could not start Chrome WebDriver ({exc}). "
            "Browser automation is disabled; falling back to programmatic "
            "embed code generation where supported.",
            file=sys.stderr,
        )
        return None


def _dismiss_consent_dialogs(driver: "webdriver.Chrome") -> None:
    """Try to dismiss common cookie/consent overlays."""
    consent_selectors = [
        'button[aria-label*="Accept"]',
        'button[aria-label*="Agree"]',
        "#accept-button",
        ".accept-cookies",
        'form[action*="consent"] button',
    ]
    for sel in consent_selectors:
        try:
            btn = driver.find_element(By.CSS_SELECTOR, sel)
            if btn.is_displayed():
                btn.click()
                time.sleep(0.5)
                return
        except Exception:
            pass


def get_embed_via_browser(
    driver: "webdriver.Chrome", url: str
) -> tuple[str | None, str | None]:
    """
    Navigate to *url* in *driver*, click Share → Embed, and return a
    ``(embed_html, duration)`` tuple.  Either element may be None on failure.

    Currently implements the YouTube share flow.  Other sites can be
    added by extending the dispatch block below.
    """
    try:
        driver.get(url)
        _dismiss_consent_dialogs(driver)

        duration = get_duration_via_browser(driver, url)

        if _is_youtube_url(url):
            embed = _youtube_share_embed(driver)
        else:
            embed = _generic_share_embed(driver)

        return embed, duration

    except Exception as exc:
        print(f"  [browser] Error fetching embed for {url}: {exc}", file=sys.stderr)
        return None, None


def _youtube_share_embed(driver: "webdriver.Chrome") -> str | None:
    """Automate the YouTube Share → Embed dialog and return the embed HTML."""
    wait = WebDriverWait(driver, 20)

    # Wait until the player area is present so buttons are loaded
    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#player")))
    except Exception:
        pass  # continue anyway

    # --- Click the Share button ---
    share_btn = None
    share_selectors = [
        (By.CSS_SELECTOR, 'button[aria-label="Share"]'),
        (By.XPATH, '//button[.//yt-formatted-string[text()="Share"]]'),
        (By.XPATH, '//ytd-button-renderer[.//yt-formatted-string[text()="Share"]]//button'),
    ]
    for by, selector in share_selectors:
        try:
            share_btn = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((by, selector))
            )
            break
        except Exception:
            pass

    if share_btn is None:
        print("  [browser] Could not find YouTube Share button.", file=sys.stderr)
        return None

    driver.execute_script("arguments[0].scrollIntoView(true);", share_btn)
    time.sleep(0.5)
    share_btn.click()

    # --- Wait for the share dialog, then click the Embed tab ---
    embed_btn = None
    embed_selectors = [
        (By.XPATH, '//yt-share-panel//button[normalize-space()="Embed"]'),
        (By.XPATH, '//ytd-unified-share-panel-renderer//button[normalize-space()="Embed"]'),
        (By.CSS_SELECTOR, 'tp-yt-paper-tab[aria-label="Embed"]'),
        (By.XPATH, '//tp-yt-paper-tab[.//div[normalize-space()="Embed"]]'),
    ]
    for by, selector in embed_selectors:
        try:
            embed_btn = WebDriverWait(driver, 8).until(
                EC.element_to_be_clickable((by, selector))
            )
            break
        except Exception:
            pass

    if embed_btn is None:
        print("  [browser] Could not find Embed tab in share dialog.", file=sys.stderr)
        return None

    embed_btn.click()
    time.sleep(1)

    # --- Extract embed code from the textarea / code panel ---
    code_selectors = [
        (By.CSS_SELECTOR, "textarea.ytd-embed-code-panel"),
        (By.CSS_SELECTOR, "#embed-code"),
        (By.XPATH, '//yt-share-panel//textarea'),
        (By.XPATH, '//div[@id="embed-code"]//textarea'),
        (By.CSS_SELECTOR, "yt-copy-code-panel textarea"),
    ]
    for by, selector in code_selectors:
        try:
            el = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((by, selector))
            )
            code = el.get_attribute("value") or el.text
            if code and "<iframe" in code:
                return code.strip()
        except Exception:
            pass

    return None


def _generic_share_embed(driver: "webdriver.Chrome") -> str | None:
    """Attempt a generic Share → Embed flow for non-YouTube pages."""
    wait = WebDriverWait(driver, 10)

    # Look for any button whose text contains "Share"
    try:
        share_btn = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, '//button[contains(translate(., "SHARE", "share"), "share")]')
            )
        )
        share_btn.click()
        time.sleep(1)
    except Exception:
        return None

    # Look for an "Embed" option
    try:
        embed_btn = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, '//button[contains(translate(., "EMBED", "embed"), "embed")]')
            )
        )
        embed_btn.click()
        time.sleep(1)
    except Exception:
        return None

    # Return any <iframe> found on the page at this point
    try:
        textarea = driver.find_element(By.CSS_SELECTOR, "textarea")
        code = textarea.get_attribute("value") or textarea.text
        if "<iframe" in code:
            return code.strip()
    except Exception:
        pass

    return None


# ---------------------------------------------------------------------------
# Spreadsheet I/O  –  Excel
# ---------------------------------------------------------------------------


def _col_index(wb_ws, column_ref: str) -> int:
    """
    Resolve *column_ref* to a 1-based column index for an openpyxl worksheet.
    Accepts an integer (as string), a letter like 'B', or a header name.

    Resolution order:
    1. Plain integer  →  used as-is (1-based)
    2. Header name match in row 1  →  that column
    3. Spreadsheet column letter(s)  (A, B, … Z, AA, …)
    """
    if column_ref.isdigit():
        return int(column_ref)

    # Header name – search row 1 first so named columns are never confused
    # with column letters (e.g. "URL" or "Embed" are valid header names).
    for cell in wb_ws[1]:
        if str(cell.value or "").strip().lower() == column_ref.strip().lower():
            return cell.column

    # Single / multi-letter column (A, B, AA …)
    if re.fullmatch(r"[A-Za-z]+", column_ref):
        return openpyxl.utils.column_index_from_string(column_ref.upper())

    raise ValueError(
        f"Cannot find column '{column_ref}' in the spreadsheet. "
        "Specify a column number, letter, or exact header name."
    )


def _ensure_header(ws, col_idx: int, header: str = "Embed Code") -> None:
    """Write a header into row 1 of *col_idx* if the cell is empty."""
    if not ws.cell(row=1, column=col_idx).value:
        ws.cell(row=1, column=col_idx, value=header)


def process_excel(
    path: str,
    url_col: str,
    embed_col: str,
    headless: bool,
    duration_col: str | None = None,
    id_col: str | None = None,
) -> None:
    if not _OPENPYXL_AVAILABLE:
        sys.exit("openpyxl is not installed.  Run: pip install -r requirements.txt")

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    url_idx = _col_index(ws, url_col)
    embed_idx = _col_index(ws, embed_col)
    _ensure_header(ws, embed_idx)

    dur_idx: int | None = None
    if duration_col:
        dur_idx = _col_index(ws, duration_col)
        _ensure_header(ws, dur_idx, header="Duration")

    item_id_idx: int | None = None
    if id_col:
        item_id_idx = _col_index(ws, id_col)
        _ensure_header(ws, item_id_idx, header="Item ID")

    driver = _create_driver(headless)
    max_row = ws.max_row
    try:
        for row_num in range(2, max_row + 1):
            url = ws.cell(row=row_num, column=url_idx).value
            if not url:
                continue
            url = str(url).strip()
            print(f"Row {row_num}: {url}")
            embed, duration = _get_embed_and_duration(driver, url)
            if embed:
                ws.cell(row=row_num, column=embed_idx, value=embed)
                print(f"  → embed code written ({len(embed)} chars)")
            else:
                print(f"  → no embed code found", file=sys.stderr)
            if dur_idx is not None:
                if duration:
                    ws.cell(row=row_num, column=dur_idx, value=duration)
                    print(f"  → duration: {duration}")
                else:
                    print(f"  → duration not found", file=sys.stderr)
            if item_id_idx is not None:
                item_id = str(uuid.uuid4())
                ws.cell(row=row_num, column=item_id_idx, value=item_id)
                print(f"  → item ID: {item_id}")
    finally:
        if driver:
            driver.quit()

    wb.save(path)
    print(f"\nSaved: {path}")


# ---------------------------------------------------------------------------
# Spreadsheet I/O  –  CSV
# ---------------------------------------------------------------------------


def _col_index_csv(headers: list[str], column_ref: str) -> int:
    """
    Resolve *column_ref* to a 0-based index for a CSV row list.
    Accepts an integer (1-based), letter, or header name.

    Resolution order:
    1. Plain integer  →  used as 1-based, converted to 0-based
    2. Header name match  →  that index
    3. Spreadsheet column letter(s)  (A, B, … AA, …)
    """
    if column_ref.isdigit():
        return int(column_ref) - 1

    # Header name – check first so that words like "URL" or "Embed" are
    # matched against actual column headers before being treated as letters.
    for i, h in enumerate(headers):
        if h.strip().lower() == column_ref.strip().lower():
            return i

    # Column letter(s) fallback
    if re.fullmatch(r"[A-Za-z]+", column_ref):
        val = 0
        for ch in column_ref.upper():
            val = val * 26 + (ord(ch) - ord("A") + 1)
        return val - 1

    raise ValueError(
        f"Cannot find column '{column_ref}' in the CSV headers. "
        "Specify a column number, letter, or exact header name."
    )


def process_csv(
    path: str,
    url_col: str,
    embed_col: str,
    headless: bool,
    duration_col: str | None = None,
    id_col: str | None = None,
) -> None:
    # Read all rows
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        print("CSV file is empty.")
        return

    headers = rows[0]
    url_idx = _col_index_csv(headers, url_col)

    # Resolve embed column; add header if needed
    try:
        embed_idx = _col_index_csv(headers, embed_col)
    except ValueError:
        embed_idx = None

    if embed_idx is None or embed_idx >= len(headers):
        embed_idx = len(headers)
        headers.append(embed_col if not embed_col.isdigit() else "Embed Code")
        for row in rows[1:]:
            while len(row) <= embed_idx:
                row.append("")

    # Resolve duration column; add header if needed
    dur_idx: int | None = None
    if duration_col:
        try:
            dur_idx = _col_index_csv(headers, duration_col)
        except ValueError:
            dur_idx = None
        if dur_idx is None or dur_idx >= len(headers):
            dur_idx = len(headers)
            headers.append(duration_col if not duration_col.isdigit() else "Duration")
            for row in rows[1:]:
                while len(row) <= dur_idx:
                    row.append("")

    # Resolve item ID column; add header if needed
    item_id_idx: int | None = None
    if id_col:
        try:
            item_id_idx = _col_index_csv(headers, id_col)
        except ValueError:
            item_id_idx = None
        if item_id_idx is None or item_id_idx >= len(headers):
            item_id_idx = len(headers)
            headers.append(id_col if not id_col.isdigit() else "Item ID")
            for row in rows[1:]:
                while len(row) <= item_id_idx:
                    row.append("")

    driver = _create_driver(headless)

    try:
        for i, row in enumerate(rows[1:], start=2):
            needed = max(
                url_idx,
                embed_idx,
                dur_idx if dur_idx is not None else 0,
                item_id_idx if item_id_idx is not None else 0,
            )
            while len(row) <= needed:
                row.append("")
            url = row[url_idx].strip()
            if not url:
                continue
            print(f"Row {i}: {url}")
            embed, duration = _get_embed_and_duration(driver, url)
            if embed:
                row[embed_idx] = embed
                print(f"  → embed code written ({len(embed)} chars)")
            else:
                print(f"  → no embed code found", file=sys.stderr)
            if dur_idx is not None:
                if duration:
                    row[dur_idx] = duration
                    print(f"  → duration: {duration}")
                else:
                    print(f"  → duration not found", file=sys.stderr)
            if item_id_idx is not None:
                item_id = str(uuid.uuid4())
                row[item_id_idx] = item_id
                print(f"  → item ID: {item_id}")
    finally:
        if driver:
            driver.quit()

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows[1:])

    print(f"\nSaved: {path}")


# ---------------------------------------------------------------------------
# Central embed-fetching logic
# ---------------------------------------------------------------------------


def _get_embed_and_duration(
    driver, url: str
) -> tuple[str | None, str | None]:
    """
    Return ``(embed_html, duration)`` for *url* using:
    1. Browser automation (if selenium is available and driver is provided)
    2. Programmatic YouTube embed code + network duration fetch as fallback
    """
    embed: str | None = None
    duration: str | None = None

    # Try browser first
    if driver:
        embed, duration = get_embed_via_browser(driver, url)

    # Programmatic fallback for YouTube embed
    if not embed:
        video_id = extract_youtube_id(url)
        if video_id:
            print("  [fallback] Using programmatic YouTube embed code.")
            embed = build_youtube_embed(video_id)

    # Network fallback for duration (only for YouTube, only if still missing)
    if duration is None and _is_youtube_url(url):
        duration = fetch_youtube_duration(url)

    return embed, duration


def _get_embed(driver, url: str) -> str | None:
    """Backward-compatible wrapper – returns embed HTML only."""
    embed, _ = _get_embed_and_duration(driver, url)
    return embed


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description=(
            "Extract embed codes (and optionally video durations) from video URLs "
            "in a spreadsheet and write them back to configurable columns."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("spreadsheet", help="Path to the .xlsx, .xls, or .csv file.")
    parser.add_argument(
        "--url-col",
        default="F",
        metavar="COL",
        help=(
            "Column containing video URLs. "
            "Accepts a 1-based number ('6'), a letter ('F'), or a header name. "
            "Default: F"
        ),
    )
    parser.add_argument(
        "--embed-col",
        default="O",
        metavar="COL",
        help=(
            "Column where embed code will be written. "
            "Same format as --url-col. "
            "Default: O"
        ),
    )
    parser.add_argument(
        "--duration-col",
        default="N",
        metavar="COL",
        help=(
            "Column where video duration will be written (e.g. '4:33'). "
            "Same format as --url-col. "
            "Set to empty string to disable duration extraction. "
            "Default: N"
        ),
    )
    parser.add_argument(
        "--id-col",
        default="P",
        metavar="COL",
        help=(
            "Column where a generated UUID will be written for each row. "
            "Same format as --url-col. "
            "Set to empty string to disable ID generation. "
            "Default: P"
        ),
    )
    parser.add_argument(
        "--no-headless",
        action="store_true",
        help="Run browser in visible (non-headless) mode. Useful for debugging.",
    )
    return parser.parse_args(argv)


def main(argv=None) -> int:
    args = _parse_args(argv)
    path = args.spreadsheet

    if not os.path.isfile(path):
        print(f"Error: file not found: {path}", file=sys.stderr)
        return 1

    headless = not args.no_headless
    duration_col = args.duration_col or None  # empty string → None (disabled)
    id_col = args.id_col or None  # empty string → None (disabled)
    ext = os.path.splitext(path)[1].lower()

    if ext in (".xlsx", ".xls"):
        if not _OPENPYXL_AVAILABLE:
            print(
                "Error: openpyxl is required for Excel files.\n"
                "Run: pip install -r requirements.txt",
                file=sys.stderr,
            )
            return 1
        process_excel(path, args.url_col, args.embed_col, headless, duration_col, id_col)

    elif ext == ".csv":
        process_csv(path, args.url_col, args.embed_col, headless, duration_col, id_col)

    else:
        print(
            f"Error: unsupported file type '{ext}'. Use .xlsx, .xls, or .csv.",
            file=sys.stderr,
        )
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
